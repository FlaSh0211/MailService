from django.shortcuts import render,redirect
from . models import Order
from openpyxl import Workbook # 엑셀을 만드는 api (엑셀 미설치 시에도 동작)
from django.core.mail import EmailMessage # 메일 전송 api
from io import BytesIO
from django.utils.safestring import mark_safe
import json
from django.core.mail import BadHeaderError, send_mail

# Create your views here.
failarray=[]
def home(request):
        
        data = Order.objects.all().order_by('company')
        data2 = set(Order.objects.values_list('company','emails'))
        counts = Order.objects.count()
        message = request.session.get('message')
        print(failarray)
        fails = list(failarray)
        failarray.clear()
        if message:
            del request.session['message']
        return render(request, 'index.html',{'data': data, 'data2':data2, 'message':mark_safe(json.dumps(message)), 'fail':mark_safe(json.dumps(str(fails))), 'count': range(counts)})


def upload(request):
        if request.method == 'POST':
            request.FILES['files'].save_to_database(
                model = Order,
                mapdict= ['company', 'product_name', 'count','emails']
        ) 
        return redirect('home')
       
        
def sendEmail(request):
        email = request.POST.get('email')
        text = Order.objects.filter(emails = email)
        content = []
        failarray.clear()
        wb = Workbook()
        ws = wb.active
        excelfile = BytesIO()

        ws['A1']= 'company'
        ws['B1']= 'product'
        ws['C1']= 'count'
        
        for i in text:
            content=[i.company,i.product_name,i.count]  
            ws.append(content)
        wb.close() 
        wb.save(excelfile)
        
        if email is not None:
            try:
                emailss = EmailMessage()
                emailss.subject = '발주 목록입니다'
                emailss.body = 'This context'
                emailss.from_email = 'tnsdnwjd33@naver.com'
                emailss.to = [email]                     
                emailss.attach('발주목록.xlsx', excelfile.getvalue(), 'application/ms-excel')
                emailss.send()
                print("보내짐")                             
                
                request.session['message'] = "파일 전송 완료"
            except:
               
                failarray.append(email)
                request.session['message'] = "파일 전송 실패"
            return redirect('home')
        
def delete(request):
        alls = Order.objects.all()
        alls.delete()
        return redirect('home')    
               
def sendAll(request):
        failcnt = 0
        failarray.clear()
        allprd = Order.objects.all()
        allemail = set(Order.objects.values_list('emails', flat=True))
        for e in allemail:
               
                wb = Workbook()
                ws = wb.active
                excelfile = BytesIO()

                ws['A1']= 'company'
                ws['B1']= 'product'
                ws['C1']= 'count'

                for i in allprd:
                    if i.emails == e:
                        content = []
                        content=[i.company, i.product_name, i.count]  
                        ws.append(content) 
                wb.close() 
                wb.save(excelfile) 
            
                if e is not None:
                    try:
                        emailss = EmailMessage()
                        emailss.subject = '발주 목록입니다'
                        emailss.body = 'This context'
                        emailss.from_email = 'tnsdnwjd33@naver.com'
                        emailss.to = [e]
                        emailss.attach('발주내역.xlsx', excelfile.getvalue(), 'application/ms-excel')
                        emailss.send()
                        print('보내짐')                         
                        if failcnt != 1:                               
                            request.session['message'] = "파일 전송 완료"
                    except:
                        failcnt = 1
                        failarray.append(e)
                        request.session['message'] = "파일 전송 실패"
                
        return redirect('home')                               
        